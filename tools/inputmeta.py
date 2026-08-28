# -*- coding: utf-8 -*-
"""Input-sheet structure: blocks, column headers, editable cells, dropdowns.

The Input sheets are three tables side by side, titled on row 6 and headed on
row 7. This reads that layout out of the workbook rather than hard-coding it,
so the app can draw each block as its own table with the workbook's own
headers, and knows which cells the user may edit and what may go in them.
"""
import re

import openpyxl
from openpyxl.utils import get_column_letter as gcl
from openpyxl.utils import column_index_from_string as cidx

HEADER_TITLE_ROW = 6
HEADER_ROW = 7
FIRST_DATA_ROW = 8

# Column roles, read off the row-7 header text. Roles drive both the alignment
# of a column and what an editable cell in it accepts.
ROLE_RULES = [
    ('sn', lambda h: h == 'S/N'),
    ('note', lambda h: h.lower().startswith('note')),
    ('pct', lambda h: '비율' in h),
    ('grade', lambda h: '난이도' in h),
    ('calc', lambda h: h == 'Factor'),
    ('sel', lambda h: h in ('입력', '선택') or h.startswith('선택')),
    ('num', lambda h: 'Quantity' in h or '참여인원' in h or 'M/H' in h),
    ('label', lambda h: 'Item' in h or True),
]

GRADES = ['상', '중', '하', 'SPI', 'N/A']


def ftext(v):
    if v is None:
        return None
    return v if isinstance(v, str) else getattr(v, 'text', None)


def is_formula(ws, col, row):
    fx = ftext(ws['%s%d' % (col, row)].value)
    return bool(fx and fx.startswith('='))


def role_of(header):
    h = str(header or '').replace('\n', ' ').strip()
    for name, test in ROLE_RULES:
        if test(h):
            return name
    return 'label'


def merged_map(ws):
    """cell coordinate -> the merged range it belongs to."""
    out = {}
    for rng in ws.merged_cells.ranges:
        for row in rng.cells:
            out['%s%d' % (gcl(row[1]), row[0])] = rng
    return out


def blocks_of(ws):
    """Row-6 merged titles give the block spans; fall back to one block."""
    out = []
    for rng in sorted(ws.merged_cells.ranges, key=lambda r: r.min_col):
        if rng.min_row == HEADER_TITLE_ROW and rng.max_row == HEADER_TITLE_ROW:
            title = ws.cell(HEADER_TITLE_ROW, rng.min_col).value
            if title:
                out.append((str(title).strip(), rng.min_col, rng.max_col))
    return out


def columns_of(ws, c0, c1):
    cols, mm = [], merged_map(ws)
    for c in range(c0, c1 + 1):
        col = gcl(c)
        raw = ws.cell(HEADER_ROW, c).value
        header = str(raw).replace('\n', ' ').strip() if raw not in (None, '') else ''
        role = None
        if not header:
            rng = mm.get('%s%d' % (col, HEADER_ROW))
            if rng is not None:
                head = ws.cell(HEADER_ROW, rng.min_col).value
                if head:
                    header, role = '', 'unit'   # tail of a merged header: the unit column
        cols.append({'c': col, 'h': header, 'role': role or role_of(header),
                     'hidden': bool(ws.column_dimensions[col].hidden)})
    return cols


# ------------------------------------------------------------------ dropdowns
def sqref_cells(sqref):
    out = set()
    for part in str(sqref).split():
        m = re.match(r'^\$?([A-Z]{1,3})\$?(\d+)(?::\$?([A-Z]{1,3})\$?(\d+))?$', part)
        if not m:
            continue
        c0, r0, c1, r1 = m.group(1), int(m.group(2)), m.group(3), m.group(4)
        c1, r1 = c1 or c0, int(r1) if r1 else r0
        for c in range(cidx(c0), cidx(c1) + 1):
            for r in range(r0, r1 + 1):
                out.add((gcl(c), r))
    return out


def dv_options(ws, formula1):
    """Resolve a list validation's source range to its non-empty values."""
    f = str(formula1 or '').strip()
    if f.startswith('"') and f.endswith('"'):
        return [x.strip() for x in f[1:-1].split(',') if x.strip()]
    m = re.match(r'^=?\$?([A-Z]{1,3})\$?(\d+):\$?([A-Z]{1,3})\$?(\d+)$', f)
    if not m:
        return []
    out = []
    for c in range(cidx(m.group(1)), cidx(m.group(3)) + 1):
        for r in range(int(m.group(2)), int(m.group(4)) + 1):
            v = ws.cell(r, c).value
            if v not in (None, ''):
                out.append(str(v).strip())
    return out


def validations(wsv):
    """cell -> option list, preferring the rule whose list holds the cell's value."""
    rules = []
    for dv in wsv.data_validations.dataValidation:
        if dv.type != 'list':
            continue
        opts = dv_options(wsv, dv.formula1)
        if len(opts) < 2:
            continue
        cells = sqref_cells(dv.sqref)
        rules.append((cells, opts))
    out = {}
    for cells, opts in sorted(rules, key=lambda x: -len(x[0])):
        for cell in cells:
            cur = wsv['%s%d' % cell].value
            got = out.get(cell)
            if got is None:
                out[cell] = opts
            elif cur is not None and str(cur).strip() in opts and str(cur).strip() not in got:
                out[cell] = opts
    return out


NUMBERED = re.compile(r'^\s*(\d+)\)')


def numbered_group(wsv, row, cols, c0, c1):
    """Options for a selector that names its choices down the rows below it.

    The Project Condition groups list '1) ...', '2) ...' in the item column and
    put the selection on the group's first row. Where the workbook's own list
    validation is broken, the group itself still says what the choices are.
    """
    label_cols = [c['c'] for c in cols if c['role'] in ('label', 'unit', 'sn')]
    out = []
    for r in range(row, row + 12):
        token = None
        for col in label_cols:
            v = wsv['%s%d' % (col, r)].value
            if v in (None, ''):
                continue
            s = str(v).strip()
            m = NUMBERED.match(s)
            if m:
                token = m.group(1) + ')'
                break
        if token is None:
            break
        if out and token == '1)':
            break
        if token in out:
            break
        out.append(token)
    return out if len(out) > 1 else []


def sibling_options(wsv, col, row, cols):
    """Choices taken from the rows that ask the same question.

    Some selectors have no usable list validation, but their siblings - the
    rows carrying the same prompt, e.g. "SPI 적용 / 미적용" - do, and the set of
    answers given down the column says what the choices are.
    """
    label_cols = [c['c'] for c in cols if c['role'] in ('label', 'unit')]

    def prompt(r):
        for c in reversed(label_cols):
            v = wsv['%s%d' % (c, r)].value
            if v not in (None, ''):
                return str(v).strip()
        return None

    want = prompt(row)
    if not want:
        return []
    seen, order = {}, []
    for r in range(FIRST_DATA_ROW, wsv.max_row + 1):
        if r == row or prompt(r) != want:
            continue
        v = wsv['%s%d' % (col, r)].value
        if not isinstance(v, str) or not v.strip() or len(v.strip()) > 12:
            continue
        s = v.strip()
        if s not in seen:
            seen[s] = 0
            order.append(s)
        seen[s] += 1
    # An answer only one sibling gives is that sibling's own oddity, not a choice.
    out = [s for s in order if seen[s] > 1] or order
    return out if len(out) > 1 else []


def cell_kind(wsf, wsv, col, row, role, dvs, colvals, cols, c0, c1):
    """What the app should draw for one cell: nothing, a number, or a list."""
    if is_formula(wsf, col, row):
        return {'k': 'calc'}
    v = wsv['%s%d' % (col, row)].value
    if role in ('sn', 'label', 'note', 'unit'):
        return None
    if v is None or v == '':
        return None
    if role in ('num', 'pct'):
        return {'k': 'num'} if isinstance(v, (int, float)) else None
    # grade / selector cells
    opts = dvs.get((col, row))
    if not opts:
        s = str(v).strip()
        if s.lower() in ('yes', 'no'):
            opts = ['Yes', 'No']
        elif role == 'grade':
            opts = [g for g in GRADES if g in colvals] or GRADES
        else:
            opts = (numbered_group(wsv, row, cols, c0, c1) or
                    sibling_options(wsv, col, row, cols))
    if opts:
        s = str(v).strip()
        if s not in opts:
            opts = [s] + opts
        return {'k': 'sel', 'o': opts}
    if isinstance(v, (int, float)):
        return {'k': 'num'}
    return None      # literal text the workbook fills in itself, e.g. 자동생성


def input_sheet(wbf, wbv, sheet):
    wsf, wsv = wbf[sheet], wbv[sheet]
    out = {'sheet': sheet, 'title': str(wsv['B1'].value or sheet).strip(), 'blocks': []}
    for title, c0, c1 in blocks_of(wsv):
        cols = columns_of(wsv, c0, c1)
        dvs = validations(wsv)
        colvals = {}
        for c in cols:
            colvals[c['c']] = {str(wsv.cell(r, cidx(c['c'])).value).strip()
                               for r in range(FIRST_DATA_ROW, wsv.max_row + 1)
                               if wsv.cell(r, cidx(c['c'])).value not in (None, '')}
        # A row belongs to the block if the block says something on it. A bare
        # S/N with nothing beside it is the workbook's spare capacity.
        keep = [c['c'] for c in cols if c['role'] in ('label', 'unit')]
        rows = []
        for r in range(FIRST_DATA_ROW, wsv.max_row + 1):
            if not any(wsv['%s%d' % (col, r)].value not in (None, '') for col in keep):
                continue
            edits = {}
            for c in cols:
                k = cell_kind(wsf, wsv, c['c'], r, c['role'], dvs, colvals[c['c']],
                              cols, c0, c1)
                if k:
                    edits[c['c']] = k
            rows.append({'r': r, 'e': edits})
        out['blocks'].append({'title': title, 'cols': cols, 'rows': rows})
    return out
