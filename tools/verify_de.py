# -*- coding: utf-8 -*-
"""Recompute OP1_CI / OP1_TEL from their formulas and diff against the workbook.

Every OP1 cell is evaluated from its own formula, with references chained back
through other OP1 cells to the Input sheets and the 산출기준 tables. Nothing is
special-cased: the tiered rows, the SUM-over-range rows and the plain
unit x qty rows all go through the same evaluator.

    python3 tools/verify_de.py <workbook.xlsx>
"""
import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))
from xlformula import compile_formula, evaluate, split_ref, Err   # noqa: E402


def ftext(v):
    if v is None:
        return None
    if isinstance(v, str):
        return v
    return getattr(v, 'text', None)


class Book:
    def __init__(self, path):
        self.f = openpyxl.load_workbook(path, data_only=False)
        self.v = openpyxl.load_workbook(path, data_only=True)
        self.cache = {}
        self.busy = set()
        self.errors = {}

    def cached(self, sheet, col, row):
        return self.v[sheet].cell(row, openpyxl.utils.column_index_from_string(col)).value

    def formula(self, sheet, col, row):
        return ftext(self.f[sheet].cell(row, openpyxl.utils.column_index_from_string(col)).value)

    def value(self, sheet, ref):
        """Recompute OP1 cells; treat everything else as a leaf (cached value)."""
        s, col, row = split_ref(ref)
        s = s or sheet
        key = (s, col, row)
        if key in self.cache:
            return self.cache[key]
        if not s.startswith('OP1') or key in self.busy:
            return self.cached(s, col, row)
        fx = self.formula(s, col, row)
        if not fx or not fx.startswith('='):
            val = self.cached(s, col, row)
        else:
            self.busy.add(key)
            try:
                val = evaluate(compile_formula(fx), lambda r: self.value(s, r))
            except Exception as e:                      # noqa: BLE001
                self.errors[key] = '%s: %s' % (type(e).__name__, e)
                val = self.cached(s, col, row)
            finally:
                self.busy.discard(key)
        self.cache[key] = val
        return val


def num(v):
    return float(v) if isinstance(v, (int, float)) else None


def main():
    path = Path(sys.argv[1])
    bk = Book(path)
    # every numeric column of the OP1 sheets
    COLS = ['G', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T',
            'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC']
    grand_total = grand_bad = 0
    for sheet in ('OP1_CI', 'OP1_TEL'):
        ws = bk.f[sheet]
        total = bad = 0
        problems = []
        for row in range(15, ws.max_row + 1):
            for col in COLS:
                fx = bk.formula(sheet, col, row)
                if not fx or not fx.startswith('='):
                    continue
                exp = num(bk.cached(sheet, col, row))
                if exp is None:
                    continue
                got = num(bk.value(sheet, '%s%d' % (col, row)))
                total += 1
                if got is None or abs(exp - got) > 0.005:
                    bad += 1
                    if len(problems) < 8:
                        problems.append((row, col, exp, got, fx[:60]))
        grand_total += total
        grand_bad += bad
        print('%-9s %5d formula cells recomputed, %d mismatched %s'
              % (sheet, total, bad, 'ALL MATCH' if bad == 0 else '<-- LOOK'))
        for p in problems:
            print('     r%-4d %-3s excel=%-12s mine=%-12s %s' % p)
    if bk.errors:
        print('\nformulas the evaluator could not handle: %d' % len(bk.errors))
        for k, v in list(bk.errors.items())[:8]:
            print('     %s!%s%d  %s' % (k[0], k[1], k[2], v))
    print('\nTOTAL %d cells, %d mismatched' % (grand_total, grand_bad))
    return 1 if grand_bad else 0


if __name__ == '__main__':
    raise SystemExit(main())
