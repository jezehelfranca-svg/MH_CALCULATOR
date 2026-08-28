# -*- coding: utf-8 -*-
"""Bundle the 실행사업 workbook's formula graph for the browser calculator.

The app recomputes the workbook rather than re-implementing it, so what ships is
the graph itself: every formula, every literal, and enough row metadata to draw
the sheets. Input literals are the editable cells; everything else derives.

    python3 tools/build_de_bundle.py <workbook.xlsx> [out.json]

Writes the bundle, plus <out>.expect.json holding the workbook's cached values
for the calculated sheets so the port can be checked against them.
"""
import json
import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

import openpyxl

import inputmeta

# Sheets the calculator needs. The OP2 report sheets are views of OP1 and are
# rendered by the app, so their formulas are not shipped.
SHEETS = ['Input_CI', 'Input_TEL', '산출기준_CI', '산출기준_TEL',
          'OP1_CI', 'OP1_TEL', 'Summary']
CALCULATED = ['OP1_CI', 'OP1_TEL', 'Summary']


def ftext(v):
    if v is None:
        return None
    return v if isinstance(v, str) else getattr(v, 'text', None)


def plain(v):
    if v is None:
        return None
    if hasattr(v, 'isoformat'):
        return str(v)
    return v


def build(src):
    wbf = openpyxl.load_workbook(src, data_only=False)
    wbv = openpyxl.load_workbook(src, data_only=True)
    cells, expect = {}, {}
    for name in SHEETS:
        f, v = wbf[name], wbv[name]
        for row in f.iter_rows():
            for c in row:
                ref = '%s!%s' % (name, c.coordinate)
                fx = ftext(c.value)
                if fx and fx.startswith('='):
                    cells[ref] = {'f': fx}
                    if name in CALCULATED:
                        got = plain(v[c.coordinate].value)
                        if isinstance(got, (int, float)):
                            expect[ref] = got
                else:
                    val = plain(v[c.coordinate].value)
                    if val is not None and val != '':
                        cells[ref] = {'v': val}
    return cells, expect


ROW_COLS = dict(sec=1, grp=2, cat=3, item=4, name=5, unit=6)


def activity_rows(wbv, sheet):
    """OP1 row structure, for drawing the sheets."""
    ws = wbv[sheet]
    out = []
    for r in range(15, ws.max_row + 1):
        name = ws.cell(r, ROW_COLS['name']).value
        unit = ws.cell(r, ROW_COLS['unit']).value
        if not name or not unit:
            continue
        out.append({
            'row': r,
            'section': ws.cell(r, ROW_COLS['sec']).value,
            'group': ws.cell(r, ROW_COLS['grp']).value,
            'category': ws.cell(r, ROW_COLS['cat']).value,
            'item': ws.cell(r, ROW_COLS['item']).value,
            'activity': str(name).strip(),
            'unit': str(unit).strip(),
        })
    return out


def main():
    src = Path(sys.argv[1])
    out = Path(sys.argv[2]) if len(sys.argv) > 2 else Path('de_bundle.json')
    cells, expect = build(src)
    wbf = openpyxl.load_workbook(src, data_only=False)
    wbv = openpyxl.load_workbook(src, data_only=True)

    bundle = {
        'source': src.name,
        'cells': cells,
        'meta': {
            'activities': {'ci': activity_rows(wbv, 'OP1_CI'),
                           'tel': activity_rows(wbv, 'OP1_TEL')},
            'inputs': {'ci': inputmeta.input_sheet(wbf, wbv, 'Input_CI'),
                       'tel': inputmeta.input_sheet(wbf, wbv, 'Input_TEL')},
        },
    }
    out.write_text(json.dumps(bundle, ensure_ascii=False, separators=(',', ':')), encoding='utf-8')
    exp = out.with_suffix('.expect.json')
    exp.write_text(json.dumps(expect, ensure_ascii=False, separators=(',', ':')), encoding='utf-8')

    print('%-22s %6d cells (%d formulas)' % (out.name, len(cells),
                                             sum(1 for c in cells.values() if 'f' in c)))
    print('%-22s %6d cached values for checking' % (exp.name, len(expect)))
    print('activities  C&I %d  TEL %d' % (len(bundle['meta']['activities']['ci']),
                                          len(bundle['meta']['activities']['tel'])))
    for part in ('ci', 'tel'):
        sheet = bundle['meta']['inputs'][part]
        print('%-11s %s' % (sheet['sheet'], '  '.join(
            '%s: %d rows, %d editable' % (
                b['title'].split('-')[-1].strip()[:28], len(b['rows']),
                sum(1 for r in b['rows'] for k in r['e'].values() if k['k'] != 'calc'))
            for b in sheet['blocks'])))
    print('%.0f KB' % (out.stat().st_size / 1024))


if __name__ == '__main__':
    main()
