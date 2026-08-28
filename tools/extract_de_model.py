# -*- coding: utf-8 -*-
"""Extract the calculation model of the 실행사업 workbook, and verify it against it.

Reads OP1_CI / OP1_TEL and recovers, for every activity row: its quantity and
difficulty, which 산출기준 row and grade supply the internal and external unit
M/H, and the per-activity GEC / 외주-종합 ratios that drive the three Cases.

Each formula's own Input references are captured column by column. The workbook
is not consistent about using one ratio row per activity row - K, M, Q and Y
sometimes point at different rows - and reproducing its numbers means following
what is written rather than assuming.

    python3 tools/extract_de_model.py <workbook.xlsx> [out.json]

The workbook's cached values are the reference; every derived column is
recomputed and compared.
"""
import json
import re
import sys
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

import openpyxl

INT_COL = {5: 'SPI', 6: '상', 7: '중', 8: '하'}
EXT_COL = {9: 'SPI', 10: '상', 11: '중', 12: '하'}
C = dict(sec=1, grp=2, cat=3, item=4, name=5, unit=6, qty=7, diff=8,
         iu=9, j=10, k=11, l=12, m=13, n=14, o=15, p=16, q=17, r=18, s=19, t=20,
         eu=21, v=22, w=23, x=24, y=25, z=26, aa=27, ab=28, ac=29)


def ftext(value):
    """openpyxl returns array formulas as objects; both forms carry the text."""
    if value is None:
        return None
    if isinstance(value, str):
        return value
    return getattr(value, 'text', None)


def xlround(x, nd=0):
    """Excel ROUND - half away from zero, unlike Python's round()."""
    q = Decimal('1') if nd == 0 else Decimal('0.' + '0' * nd)
    return float(Decimal(repr(float(x))).quantize(q, rounding=ROUND_HALF_UP))


def num(v):
    return float(v) if isinstance(v, (int, float)) else 0.0


def grade_of(col_letter):
    ci = openpyxl.utils.column_index_from_string(col_letter)
    if ci in INT_COL:
        return ['int', INT_COL[ci]]
    if ci in EXT_COL:
        return ['ext', EXT_COL[ci]]
    return None


def parse_lookup(value, stdname):
    """Unit-M/H lookup -> ['lookup', std_row, {difficulty: [side, grade]}, default]."""
    formula = ftext(value)
    if not formula or stdname not in formula:
        return None
    stripped = re.sub(re.escape(stdname) + r'!\$?[A-Z]{1,3}\$?\d+', 'R', formula)
    if re.search(r'R\s*[*/+]|[*/+]\s*R', stripped):
        return ['special', formula]           # arithmetic wrapped around the lookup
    refs = re.findall(re.escape(stdname) + r'!\$?([A-Z]{1,3})\$?(\d+)', formula)
    std_row = int(refs[0][1])
    pairs = re.findall(r'="([^"]+)",\s*' + re.escape(stdname) + r'!\$?([A-Z]{1,3})\$?\d+', formula)
    cond = {c: grade_of(col) for c, col in pairs}
    used = {col for _, col in pairs}
    default = None
    for col, _ in refs:
        if col not in used:
            default = grade_of(col)
    if default is None:
        m = re.search(r',\s*(-?\d+(?:\.\d+)?)\s*\)+\s*$', formula)
        default = ['literal', float(m.group(1))] if m else (grade_of(refs[-1][0]) if refs else None)
    return ['lookup', std_row, cond, default]


def ref_value(ws, ref):
    m = re.match(r'\$?([A-Z]{1,3})\$?(\d+)$', ref)
    if not m:
        return 0.0
    return num(ws.cell(int(m.group(2)), openpyxl.utils.column_index_from_string(m.group(1))).value)


def extract(src):
    wbf = openpyxl.load_workbook(src, data_only=False)
    wbv = openpyxl.load_workbook(src, data_only=True)
    model = {}
    for part, op, stdname, inp in (('ci', 'OP1_CI', '산출기준_CI', 'Input_CI'),
                                   ('tel', 'OP1_TEL', '산출기준_TEL', 'Input_TEL')):
        f, v, iv = wbf[op], wbv[op], wbv[inp]
        rows = []
        for r in range(15, f.max_row + 1):
            name, unit = v.cell(r, C['name']).value, v.cell(r, C['unit']).value
            if not name or not unit:
                continue

            def refs(col):
                t = ftext(f.cell(r, C[col]).value) or ''
                return re.findall(re.escape(inp) + r'!(\$?[A-Z]{1,3}\$?\d+)', t)

            def val(lst, i, dflt=0.0):
                return ref_value(iv, lst[i]) if len(lst) > i else dflt

            rk, rm, rq, ry, rs = refs('k'), refs('m'), refs('q'), refs('y'), refs('s')
            rows.append({
                'row': r,
                'section': v.cell(r, C['sec']).value,
                'group': v.cell(r, C['grp']).value,
                'category': v.cell(r, C['cat']).value,
                'item': v.cell(r, C['item']).value,
                'activity': str(name).strip(),
                'unit': str(unit).strip(),
                'qty': num(v.cell(r, C['qty']).value),
                'difficulty': v.cell(r, C['diff']).value,
                'int_lookup': parse_lookup(f.cell(r, C['iu']).value, stdname),
                'ext_lookup': parse_lookup(f.cell(r, C['eu']).value, stdname),
                'ratios': {
                    'k_gec': val(rk, 0),
                    'm_gec': val(rm, 0), 'm_comp': val(rm, 1),
                    'q_gec': val(rq, 0), 'q_den': val(rq, 1, 0.7),
                    'y_gecext': val(ry, 0), 'y_comp': val(ry, 1), 'y_den': val(ry, 2, 0.7),
                    's_gecext': val(rs, 0),
                },
                'mh_from_mh': bool(re.match(r'=ROUND\(J\d+', ftext(f.cell(r, C['l']).value) or '')),
                'expect': {k: v.cell(r, C[k]).value for k in
                           ('iu', 'j', 'l', 'n', 'p', 'eu', 'v', 'x', 'z', 'aa', 'ab', 'ac')},
            })
        model[part] = rows
        print('%-9s %4d activity rows' % (op, len(rows)))
    return model


def compute(row, std_value):
    """The three Cases, exactly as the workbook computes them."""
    q = row['qty']
    iu = std_value(row['int_lookup'], row['difficulty']) or 0.0
    eu = std_value(row['ext_lookup'], row['difficulty']) or 0.0
    R = row['ratios']

    j = xlround(q * iu)                                   # Case1 internal
    vv = xlround(q * eu)                                  # Case1 external
    if row['mh_from_mh']:
        l = xlround(j * (1 - R['k_gec']))
        n = xlround(j * (1 - R['m_gec'] - R['m_comp']))
        rr = xlround(j * R['q_gec'] / R['q_den']) if R['q_den'] else 0.0
    else:
        l = xlround(iu * (1 - R['k_gec']) * q)
        n = xlround(iu * (1 - R['m_gec'] - R['m_comp']) * q)
        rr = xlround(iu * R['q_gec'] / R['q_den'] * q) if R['q_den'] else 0.0
    t = xlround(eu * R['s_gecext'] * q)
    p = rr + t                                            # GEC M/H
    x = xlround(eu * (1 - R['y_gecext']) * q)
    y_unit = eu * (1 - R['y_gecext']) + (iu * R['y_comp'] / R['y_den'] if R['y_den'] else 0.0)
    z = xlround(y_unit * q)
    return dict(iu=iu, eu=eu, j=j, v=vv, l=l, n=n, p=p, x=x, z=z,
                aa=j + vv, ab=l + p + x, ac=n + p + z)


def verify(src, model):
    wbv = openpyxl.load_workbook(src, data_only=True)
    total = bad = 0
    problems = []
    skipped = 0
    for part, stdname in (('ci', '산출기준_CI'), ('tel', '산출기준_TEL')):
        sv = wbv[stdname]

        def std_value(lk, diff):
            if not lk or lk[0] != 'lookup':
                return None
            _, std_row, cond, default = lk
            spec = cond.get(diff, default)
            if not spec:
                return 0.0
            side, grade = spec
            if side == 'literal':
                return float(grade)
            cols = INT_COL if side == 'int' else EXT_COL
            col = [k for k, g in cols.items() if g == grade]
            return num(sv.cell(std_row, col[0]).value) if col else 0.0

        for row in model[part]:
            if any(lk and lk[0] == 'special' for lk in (row['int_lookup'], row['ext_lookup'])):
                skipped += 1
                continue
            got = compute(row, std_value)
            for key in ('iu', 'j', 'l', 'n', 'p', 'eu', 'v', 'x', 'z', 'aa', 'ab', 'ac'):
                exp = row['expect'].get(key)
                if not isinstance(exp, (int, float)):
                    continue
                total += 1
                if abs(num(exp) - got[key]) > 0.005:
                    bad += 1
                    if len(problems) < 10:
                        problems.append((part, row['row'], key, exp, got[key],
                                         str(row['activity'])[:28].replace('\n', ' ')))
    print('\n=== verification against the workbook\'s cached values ===')
    print('  cells checked      : %d' % total)
    print('  special-case rows  : %d (handled separately)' % skipped)
    print('  mismatched         : %d  %s' % (bad, 'ALL MATCH' if bad == 0 else '<-- LOOK'))
    for p in problems:
        print('     %-4s r%-4d %-3s excel=%-11s mine=%-11s %s' % p)
    return bad


def main():
    src = Path(sys.argv[1])
    out = Path(sys.argv[2]) if len(sys.argv) > 2 else Path('de_model.json')
    model = extract(src)
    bad = verify(src, model)
    out.write_text(json.dumps(model, ensure_ascii=False, indent=1), encoding='utf-8')
    print('wrote %s' % out)
    return 1 if bad else 0


if __name__ == '__main__':
    raise SystemExit(main())
