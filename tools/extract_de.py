# -*- coding: utf-8 -*-
"""Extract the 실행사업 (detail engineering) M/H standard from the source workbook.

The workbook stores BOTH outsourcing versions inside each standards cell as
    =IF(A5="일반 Ver.", <standard>, <minimized>)
with A5 driven by Input_CI!E3. So one file yields both unit sets, and
MH_Calculator_Propoal.xlsx is the same book with that switch flipped.

    python3 tools/extract_de.py <workbook.xlsx> [out.json]
"""
import json
import re
import sys
from pathlib import Path

import openpyxl

VER_IF = re.compile(r'^=IF\(A5="일반 Ver\.",\s*(?P<std>[^,]+?),\s*(?P<min>[^)]+?)\)$')
GRADES = ['SPI', '상', '중', '하']


def cell_versions(formula, cached):
    """Return (standard, minimized) values for one unit-M/H cell."""
    if isinstance(formula, str) and formula.startswith('='):
        m = VER_IF.match(formula.strip())
        if m:
            def lit(t):
                t = t.strip()
                try:
                    return float(t)
                except ValueError:
                    return t.strip('"')
            return lit(m.group('std')), lit(m.group('min'))
        # a formula we do not model - fall back to the cached value for both
        return cached, cached
    return formula, formula


def to_num(v):
    if isinstance(v, (int, float)):
        return float(v)
    return 0.0


def extract_standard(wbf, wbv, sheet):
    """One 산출기준 sheet -> ordered activity records carrying both versions."""
    f, v = wbf[sheet], wbv[sheet]
    out, section, group = [], None, None
    for r in range(1, f.max_row + 1):
        g = lambda c: v.cell(r, c).value
        a, b, c3, unit = g(1), g(2), g(3), g(4)

        if isinstance(a, str) and len(str(a).strip()) <= 2 and b and not unit:
            section = '%s %s' % (str(a).strip(), str(b).strip())
            group = None
            continue
        if isinstance(a, (int, float)) and b and not unit:
            group = '%s. %s' % (int(a), str(b).strip())
            continue
        if not unit:
            continue

        std_vals, min_vals = {}, {}
        any_num = False
        for i, grade in enumerate(GRADES):          # internal: cols E..H (5-8)
            s, m = cell_versions(f.cell(r, 5 + i).value, g(5 + i))
            std_vals['int_' + grade], min_vals['int_' + grade] = s, m
            any_num |= isinstance(s, (int, float))
        for i, grade in enumerate(GRADES):          # external: cols I..L (9-12)
            s, m = cell_versions(f.cell(r, 9 + i).value, g(9 + i))
            std_vals['ext_' + grade], min_vals['ext_' + grade] = s, m
            any_num |= isinstance(s, (int, float))
        if not any_num:
            continue

        name = c3 if c3 else b
        out.append({
            'row': r,
            'section': section,
            'group': group,
            'activity': str(name).strip(),
            'unit': str(unit).strip(),
            'standard': {k: to_num(x) for k, x in std_vals.items()},
            'minimized': {k: to_num(x) for k, x in min_vals.items()},
            'guide': str(g(14) or g(13) or '').strip(),
        })
    return out


def main():
    src = Path(sys.argv[1])
    dst = Path(sys.argv[2]) if len(sys.argv) > 2 else Path('de_data.json')
    wbf = openpyxl.load_workbook(src, data_only=False)
    wbv = openpyxl.load_workbook(src, data_only=True)

    data = {}
    for part, sheet in (('ci', '산출기준_CI'), ('tel', '산출기준_TEL')):
        rows = extract_standard(wbf, wbv, sheet)
        data[part] = rows
        differing = sum(1 for x in rows if x['standard'] != x['minimized'])
        print('%-14s %4d activities   %3d differ between the two versions'
              % (sheet, len(rows), differing))

    dst.write_text(json.dumps(data, ensure_ascii=False, indent=1), encoding='utf-8')
    print('wrote %s' % dst)


if __name__ == '__main__':
    main()
